#!/usr/bin/env python3
# -*- coding：utf-8 -*-

import os
import shutil 

INPUT_FILE_BASE_PATH = 'E:\\PyWork\\BaseCore\\testdir\\'
EXCEL_FILENAME = 'sample.xlsx'
FILE_NAME = r"{}".format(INPUT_FILE_BASE_PATH) + "{}".format(EXCEL_FILENAME)

from openpyxl import load_workbook  #必须的 打开Excel文件用的
from openpyxl import Workbook  #必须的 创建Excel文件缓冲用的

def createExcel():
	wb = Workbook()
	ws = wb.active
	# 数据可以直接分配到单元格中
	ws['A1'] = 42
	# 可以附加行，从第一列开始附加
	ws.append([1, 2, 3])
	# Python 类型会被自动转换
	import datetime
	ws['A3'] = datetime.datetime.now().strftime("%Y-%m-%d")
	ws.title = "Base"

	# worksheet
	ws1 = wb.create_sheet("Mysheet") #插入到最后(default)
	ws2 = wb.create_sheet("Mysheet", 0) #插入到最开始的位置

	ws1.title = "New Title"
	ws1['A1'] = 'New11'
	ws1.sheet_properties.tabColor = "1072BA"
	
	ws2.title = "New Title2"
	ws2['A2'] = 'New22'
	# 使用公式
	ws2["A3"] = "=SUM(1, 1)"

	for sheet in wb:
		print(sheet.title)

	# 保存文件
	wb.save(EXCEL_FILENAME)


#from openpyxl.writer.excel import ExcelWriter  #向Excel写数据需要导入的模块
from openpyxl.reader.excel import load_workbook  #从Excel读数据需要导入的模块
import json

def readExcel():
	wb = load_workbook(filename=FILE_NAME)
	#该函数返回sheet 名称列表
	print ( "Worksheet name(s):", wb.sheetnames)
	sheetnames = wb.sheetnames
	ws = wb[sheetnames[1]] #通过sheet名获取sheet的 流对象
	print ("Work Sheet Titile:",ws.title)
	print ("Work Sheet Rows:",ws.max_row) #表行数
	print ("Work Sheet Cols:",ws.max_column) #表列数

def readExcel2():
	wb = load_workbook(filename=FILE_NAME)
	print ( "Worksheet name(s):", wb.sheetnames[1])
	ws = wb[wb.sheetnames[1]]	
	# 建立存储数据的字典
	data_dic = {}
	for rx in range(1,ws.max_row + 1):
		temp_list = []
		pid = rx
		w1 = ws.cell(row=rx, column=1).value
		w2 = ws.cell(row=rx, column=2).value
		w3 = ws.cell(row=rx, column=3).value
		w4 = ws.cell(row=rx, column=4).value
		w5 = ws.cell(row=rx, column=5).value
		temp_list = [w1, w2, w3, w4,w5]
		data_dic[pid] = temp_list
	# 打印字典数据个数
	print ('Total:%d' % len(data_dic))
	print (json.dumps(data_dic,ensure_ascii=False))


if __name__ == '__main__':
	readExcel2()